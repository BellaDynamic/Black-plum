#!/usr/bin/env python3
"""
Street Legends 30-Day Action Plan Tracking Spreadsheet
Professional Excel workbook with comprehensive stakeholder tracking
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Theme: Elegant Black
THEME = {
    'primary': '2D2D2D',
    'light': 'E5E5E5',
    'accent': '2D2D2D',
    'text': '000000',
    'white': 'FFFFFF'
}

SERIF_FONT = 'Georgia'
SANS_FONT = 'Calibri'

def create_30day_tracker():
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    wb.remove(wb.active)
    
    # Create sheets
    overview = wb.create_sheet("Overview", 0)
    week1 = wb.create_sheet("Week 1 (Feb 1-7)", 1)
    week2 = wb.create_sheet("Week 2 (Feb 8-14)", 2)
    week3 = wb.create_sheet("Week 3 (Feb 15-21)", 3)
    week4 = wb.create_sheet("Week 4 (Feb 22-28)", 4)
    stakeholders = wb.create_sheet("Stakeholder Database", 5)
    contracts = wb.create_sheet("Contract Status", 6)
    
    # Setup each sheet
    setup_overview(overview)
    setup_week1(week1)
    setup_week2(week2)
    setup_week3(week3)
    setup_week4(week4)
    setup_stakeholders(stakeholders)
    setup_contracts(contracts)
    
    # Save workbook
    wb.save('/home/ubuntu/cohort-framework/Street-Legends-30Day-Tracker.xlsx')
    print("✅ Created: Street-Legends-30Day-Tracker.xlsx")

def setup_overview(ws):
    """Create overview dashboard"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    # Title
    ws['B2'] = "STREET LEGENDS: 30-DAY ACTION PLAN TRACKER"
    ws['B2'].font = Font(name=SERIF_FONT, size=18, bold=True, color=THEME['accent'])
    
    ws['B3'] = "Phase 0: Assess/Pre-Event | February 1 - March 1, 2026"
    ws['B3'].font = Font(name=SANS_FONT, size=11, italic=True, color='666666')
    
    # Mission statement
    ws['B5'] = "MISSION"
    ws['B5'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['accent'])
    
    ws['B6'] = "Secure all foundational event and stakeholder contracts to green-light the Street Legends travel docu-series for production."
    ws['B6'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[6].height = 30
    
    # Weekly objectives
    ws['B9'] = "WEEKLY OBJECTIVES"
    ws['B9'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['accent'])
    
    objectives = [
        ("Week 1 (Feb 1-7)", "Stakeholder Alignment & Onboarding", "Import contacts to Align, send Reality Check messages, schedule discovery calls"),
        ("Week 2 (Feb 8-14)", "Contract Distribution & Negotiation", "Distribute contracts, negotiate terms, define Super Circus operations"),
        ("Week 3 (Feb 15-21)", "Finalize Agreements & Secure Commitments", "Finalize contracts, secure commitments, create Right of Use agreements"),
        ("Week 4 (Feb 22-28)", "Execute Contracts & Prepare for State Funding", "Obtain signatures, collect capital, prepare state funding requests")
    ]
    
    row = 11
    for week, objective, details in objectives:
        ws.cell(row=row, column=2, value=week).font = Font(name=SANS_FONT, size=11, bold=True)
        ws.cell(row=row, column=3, value=objective).font = Font(name=SANS_FONT, size=11, bold=True, color=THEME['accent'])
        ws.cell(row=row+1, column=3, value=details).font = Font(name=SANS_FONT, size=10, italic=True)
        ws.cell(row=row+1, column=3).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row+1].height = 30
        row += 3
    
    # Key stakeholder categories
    ws['B23'] = "KEY STAKEHOLDER CATEGORIES"
    ws['B23'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['accent'])
    
    categories = [
        "• NHRA Events - Venue contracts, filming permissions, event dates",
        "• NASCAR Events - Venue contracts, filming permissions, event dates",
        "• Super Circus - Master service agreement (state-by-state operations)",
        "• Auto Industry Sponsors - Contribution agreements, branding rights",
        "• Event Crew - Service agreements for live event execution",
        "• Z Crew (Content Capture) - Production service agreements",
        "• Venues - Filming permissions, access agreements, insurance"
    ]
    
    for i, category in enumerate(categories, start=25):
        ws.cell(row=i, column=2, value=category).font = Font(name=SANS_FONT, size=10)
    
    # Success criteria
    ws['B34'] = "SUCCESS CRITERIA (March 1 Deadline)"
    ws['B34'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['accent'])
    
    criteria = [
        "✓ All 50+ stakeholders onboarded to Align platform",
        "✓ Master Alignment Agreement (MAA) executed with all key parties",
        "✓ Super Circus state-by-state operations defined and contracted",
        "✓ All venue contracts signed with filming permissions",
        "✓ Event Crew and Z Crew service agreements executed",
        "✓ Sponsor contribution agreements signed with Right of Use",
        "✓ Initial capital contributions collected",
        "✓ State/local funding requests prepared and submitted",
        "✓ Project GREEN-LIT for production planning"
    ]
    
    for i, criterion in enumerate(criteria, start=36):
        ws.cell(row=i, column=2, value=criterion).font = Font(name=SANS_FONT, size=10)
    
    # Navigation
    ws['B47'] = "NAVIGATION"
    ws['B47'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['accent'])
    
    sheets = ["Week 1 (Feb 1-7)", "Week 2 (Feb 8-14)", "Week 3 (Feb 15-21)", "Week 4 (Feb 22-28)", "Stakeholder Database", "Contract Status"]
    for i, sheet_name in enumerate(sheets, start=49):
        cell = ws.cell(row=i, column=2, value=f"→ {sheet_name}")
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = Font(name=SANS_FONT, size=10, color=THEME['accent'], underline='single')
    
    # Footer
    ws['B57'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws['B57'].font = Font(name=SANS_FONT, size=9, italic=True, color='999999')
    
    # Column widths
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60

def setup_week1(ws):
    """Week 1: Stakeholder Alignment & Onboarding"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "WEEK 1: Stakeholder Alignment & Onboarding"
    ws['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['accent'])
    ws['B3'] = "February 1-7, 2026"
    ws['B3'].font = Font(name=SANS_FONT, size=11, italic=True)
    
    # Create task tracking table
    headers = ["Task", "Owner", "Due Date", "Status", "Notes"]
    tasks = [
        ("Import all 50+ contacts to Align platform", "Brandy", "Feb 2", "Not Started", "Create 'Street Legends' room"),
        ("Tag contacts with provisional roles", "Brandy", "Feb 2", "Not Started", "Sponsor/Venue/Service Provider"),
        ("Send Reality Check message to all stakeholders", "Brandy", "Feb 3", "Not Started", "Use template from docs"),
        ("Schedule discovery call: Super Circus", "Brandy", "Feb 4", "Not Started", "Priority #1"),
        ("Schedule discovery call: NHRA venues", "Brandy", "Feb 4", "Not Started", "Priority #2"),
        ("Schedule discovery call: NASCAR venues", "Brandy", "Feb 4", "Not Started", "Priority #3"),
        ("Schedule discovery call: Auto Industry sponsors", "Brandy", "Feb 5", "Not Started", "Identify top 5"),
        ("Prepare contract templates for review", "Brandy", "Feb 6", "Not Started", "5 templates ready"),
        ("Create Jimmy McNickle private Align room", "Brandy", "Feb 2", "Not Started", "Separate from main project"),
        ("Hold initial evaluation meetings (top 10)", "Brandy", "Feb 5-7", "Not Started", "Determine pathways")
    ]
    
    row = 6
    # Headers
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name=SANS_FONT, size=11, bold=True, color=THEME['white'])
        cell.fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin', color='000000'))
    
    # Tasks
    for task, owner, due, status, notes in tasks:
        row += 1
        ws.cell(row=row, column=2, value=task).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=3, value=owner).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=4, value=due).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=5, value=status).font = Font(name=SANS_FONT, size=10, color='FF0000')
        ws.cell(row=row, column=6, value=notes).font = Font(name=SANS_FONT, size=9, italic=True)
    
    # Column widths
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    # Back to overview link
    ws[f'B{row+3}'] = "← Back to Overview"
    ws[f'B{row+3}'].hyperlink = "#'Overview'!A1"
    ws[f'B{row+3}'].font = Font(name=SANS_FONT, size=10, color=THEME['accent'], underline='single')

def setup_week2(ws):
    """Week 2: Contract Distribution & Negotiation"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "WEEK 2: Contract Distribution & Negotiation"
    ws['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['accent'])
    ws['B3'] = "February 8-14, 2026"
    ws['B3'].font = Font(name=SANS_FONT, size=11, italic=True)
    
    headers = ["Task", "Owner", "Due Date", "Status", "Notes"]
    tasks = [
        ("Distribute Venue Agreement to NHRA contacts", "Brandy", "Feb 8", "Not Started", "Include filming permissions"),
        ("Distribute Venue Agreement to NASCAR contacts", "Brandy", "Feb 8", "Not Started", "Include filming permissions"),
        ("Distribute Super Circus Master Service Agreement", "Brandy", "Feb 9", "Not Started", "State-by-state operations"),
        ("Negotiate Super Circus load-in/load-out terms", "Brandy", "Feb 10", "Not Started", "Critical for logistics"),
        ("Distribute Sponsor/Contributor Agreements", "Brandy", "Feb 11", "Not Started", "Right of Use included"),
        ("Distribute Event Crew Service Agreements", "Brandy", "Feb 12", "Not Started", "Define scope of work"),
        ("Distribute Z Crew Service Agreements", "Brandy", "Feb 12", "Not Started", "Content capture specs"),
        ("Negotiate sponsor branding rights", "Brandy", "Feb 13", "Not Started", "Logo placement, credits"),
        ("Define Super Circus concession operations", "Brandy", "Feb 14", "Not Started", "Revenue sharing model"),
        ("Review all contract feedback", "Brandy", "Feb 14", "Not Started", "Prepare for Week 3 finalization")
    ]
    
    row = 6
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name=SANS_FONT, size=11, bold=True, color=THEME['white'])
        cell.fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin', color='000000'))
    
    for task, owner, due, status, notes in tasks:
        row += 1
        ws.cell(row=row, column=2, value=task).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=3, value=owner).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=4, value=due).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=5, value=status).font = Font(name=SANS_FONT, size=10, color='FF0000')
        ws.cell(row=row, column=6, value=notes).font = Font(name=SANS_FONT, size=9, italic=True)
    
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    ws[f'B{row+3}'] = "← Back to Overview"
    ws[f'B{row+3}'].hyperlink = "#'Overview'!A1"
    ws[f'B{row+3}'].font = Font(name=SANS_FONT, size=10, color=THEME['accent'], underline='single')

def setup_week3(ws):
    """Week 3: Finalize Agreements & Secure Commitments"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "WEEK 3: Finalize Agreements & Secure Commitments"
    ws['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['accent'])
    ws['B3'] = "February 15-21, 2026"
    ws['B3'].font = Font(name=SANS_FONT, size=11, italic=True)
    
    headers = ["Task", "Owner", "Due Date", "Status", "Notes"]
    tasks = [
        ("Finalize all contract terms", "Brandy", "Feb 17", "Not Started", "All parties aligned"),
        ("Secure verbal commitments from all stakeholders", "Brandy", "Feb 17", "Not Started", "Document in Align"),
        ("Create Right of Use Agreements for contributors", "Brandy", "Feb 18", "Not Started", "Link funds to project"),
        ("Lock down filming permissions with all venues", "Brandy", "Feb 18", "Not Started", "Z Crew access confirmed"),
        ("Finalize Super Circus state-by-state operations", "Brandy", "Feb 19", "Not Started", "All states defined"),
        ("Confirm Event Crew availability and rates", "Brandy", "Feb 19", "Not Started", "Lock in team"),
        ("Confirm Z Crew availability and rates", "Brandy", "Feb 20", "Not Started", "Lock in team"),
        ("Prepare Master Alignment Agreement (MAA)", "Brandy", "Feb 21", "Not Started", "Incorporate all sub-agreements"),
        ("Review all insurance requirements", "Brandy", "Feb 21", "Not Started", "Venues, crews, events"),
        ("Prepare Week 4 signature packets", "Brandy", "Feb 21", "Not Started", "Ready for execution")
    ]
    
    row = 6
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name=SANS_FONT, size=11, bold=True, color=THEME['white'])
        cell.fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin', color='000000'))
    
    for task, owner, due, status, notes in tasks:
        row += 1
        ws.cell(row=row, column=2, value=task).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=3, value=owner).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=4, value=due).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=5, value=status).font = Font(name=SANS_FONT, size=10, color='FF0000')
        ws.cell(row=row, column=6, value=notes).font = Font(name=SANS_FONT, size=9, italic=True)
    
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    ws[f'B{row+3}'] = "← Back to Overview"
    ws[f'B{row+3}'].hyperlink = "#'Overview'!A1"
    ws[f'B{row+3}'].font = Font(name=SANS_FONT, size=10, color=THEME['accent'], underline='single')

def setup_week4(ws):
    """Week 4: Execute Contracts & Prepare for State Funding"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "WEEK 4: Execute Contracts & Prepare for State Funding"
    ws['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['accent'])
    ws['B3'] = "February 22-28, 2026"
    ws['B3'].font = Font(name=SANS_FONT, size=11, italic=True)
    
    headers = ["Task", "Owner", "Due Date", "Status", "Notes"]
    tasks = [
        ("Obtain signatures on all venue contracts", "Brandy", "Feb 24", "Not Started", "Digital via Align/Beneflts"),
        ("Obtain signatures on Super Circus agreement", "Brandy", "Feb 24", "Not Started", "Master service agreement"),
        ("Obtain signatures on Event Crew agreements", "Brandy", "Feb 25", "Not Started", "All crew members"),
        ("Obtain signatures on Z Crew agreements", "Brandy", "Feb 25", "Not Started", "All production crew"),
        ("Obtain signatures on Sponsor agreements", "Brandy", "Feb 26", "Not Started", "All contributors"),
        ("Collect initial capital contributions", "Brandy", "Feb 26", "Not Started", "Per payment schedules"),
        ("Execute Master Alignment Agreement (MAA)", "Brandy", "Feb 27", "Not Started", "All parties sign"),
        ("Package contracts for state funding requests", "Brandy", "Feb 27", "Not Started", "Economic dev, tourism"),
        ("Submit state/local funding requests", "Brandy", "Feb 28", "Not Started", "All applicable states"),
        ("PROJECT GREEN-LIT: Begin production planning", "Brandy", "Mar 1", "Not Started", "Enter Phase 1: Assess/Pre-GRID")
    ]
    
    row = 6
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name=SANS_FONT, size=11, bold=True, color=THEME['white'])
        cell.fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin', color='000000'))
    
    for task, owner, due, status, notes in tasks:
        row += 1
        ws.cell(row=row, column=2, value=task).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=3, value=owner).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=4, value=due).font = Font(name=SANS_FONT, size=10)
        ws.cell(row=row, column=5, value=status).font = Font(name=SANS_FONT, size=10, color='FF0000')
        ws.cell(row=row, column=6, value=notes).font = Font(name=SANS_FONT, size=9, italic=True)
    
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    ws[f'B{row+3}'] = "← Back to Overview"
    ws[f'B{row+3}'].hyperlink = "#'Overview'!A1"
    ws[f'B{row+3}'].font = Font(name=SANS_FONT, size=10, color=THEME['accent'], underline='single')

def setup_stakeholders(ws):
    """Stakeholder database"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "STAKEHOLDER DATABASE"
    ws['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['accent'])
    ws['B3'] = "Track all 50+ Street Legends stakeholders"
    ws['B3'].font = Font(name=SANS_FONT, size=11, italic=True)
    
    headers = ["Name", "Organization", "Category", "Role", "Contact", "Align Status", "Priority", "Notes"]
    
    row = 6
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name=SANS_FONT, size=11, bold=True, color=THEME['white'])
        cell.fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin', color='000000'))
    
    # Sample entries
    samples = [
        ("", "", "NHRA", "Venue Contact", "", "Not Invited", "High", ""),
        ("", "", "NASCAR", "Venue Contact", "", "Not Invited", "High", ""),
        ("", "", "Super Circus", "Service Provider", "", "Not Invited", "Critical", "Master agreement required"),
        ("", "", "Auto Industry", "Sponsor", "", "Not Invited", "High", ""),
        ("", "", "Event Crew", "Service Provider", "", "Not Invited", "Medium", ""),
        ("", "", "Z Crew", "Production", "", "Not Invited", "Medium", "Content capture"),
    ]
    
    for sample in samples:
        row += 1
        for col, value in enumerate(sample, start=2):
            ws.cell(row=row, column=col, value=value).font = Font(name=SANS_FONT, size=10)
    
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 30
    
    ws[f'B{row+3}'] = "← Back to Overview"
    ws[f'B{row+3}'].hyperlink = "#'Overview'!A1"
    ws[f'B{row+3}'].font = Font(name=SANS_FONT, size=10, color=THEME['accent'], underline='single')

def setup_contracts(ws):
    """Contract status tracking"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    
    ws['B2'] = "CONTRACT STATUS TRACKER"
    ws['B2'].font = Font(name=SERIF_FONT, size=16, bold=True, color=THEME['accent'])
    ws['B3'] = "Track all contract executions for Master Alignment Agreement"
    ws['B3'].font = Font(name=SANS_FONT, size=11, italic=True)
    
    headers = ["Stakeholder", "Contract Type", "Sent Date", "Reviewed", "Negotiated", "Signed Date", "Status", "Notes"]
    
    row = 6
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name=SANS_FONT, size=11, bold=True, color=THEME['white'])
        cell.fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin', color='000000'))
    
    # Sample entries
    samples = [
        ("", "Venue Agreement", "", "No", "No", "", "Not Started", "NHRA venue"),
        ("", "Venue Agreement", "", "No", "No", "", "Not Started", "NASCAR venue"),
        ("", "Super Circus Master", "", "No", "No", "", "Not Started", "State-by-state ops"),
        ("", "Sponsor Agreement", "", "No", "No", "", "Not Started", "Right of Use"),
        ("", "Event Crew Service", "", "No", "No", "", "Not Started", ""),
        ("", "Z Crew Service", "", "No", "No", "", "Not Started", "Content capture"),
    ]
    
    for sample in samples:
        row += 1
        for col, value in enumerate(sample, start=2):
            ws.cell(row=row, column=col, value=value).font = Font(name=SANS_FONT, size=10)
    
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    
    ws[f'B{row+3}'] = "← Back to Overview"
    ws[f'B{row+3}'].hyperlink = "#'Overview'!A1"
    ws[f'B{row+3}'].font = Font(name=SANS_FONT, size=10, color=THEME['accent'], underline='single')

if __name__ == "__main__":
    create_30day_tracker()
